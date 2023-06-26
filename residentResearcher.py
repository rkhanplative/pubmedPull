from Bio import Entrez
import ssl
import openpyxl
from datetime import datetime
from datetime import timedelta
from datetime import date 

ssl._create_default_https_context = ssl._create_unverified_context

# Provide your email address to the Entrez API
Entrez.email = "ruk6@scarletmail.rutgers.edu"

file_name = 'Table 8c Part III Current or Former residents.xlsx'
input_sheet_name = 'Sheet2'
output_sheet_name = 'Sheet3'

# Load the workbook
workbook = openpyxl.load_workbook(file_name)

# Get the sheet by name
input_sheet = workbook[input_sheet_name]
output_sheet = workbook[output_sheet_name]

# Define an empty list to store the data
data = []

# Loop through each row in the sheet and append the values to the list
for row in input_sheet.iter_rows(values_only=True):
    data.append(row)

def search_pub_med(name,start_date,end_date,position):
    # Specify the search term
    search_term = f'("{start_date}"[Date - Publication] : "{end_date}"[Date - Publication]) AND ({name}[{position}] AND ({name}[Exact name])'

    # Use the Entrez API to search PubMed for articles matching the search term
    handle = Entrez.esearch(db="pubmed", term=search_term)
    record = Entrez.read(handle)

    # Extract the number of articles found from the search results
    num_articles = record["Count"]
    return num_articles


def get_data():
    #Define an empty list to store the data
    output = []
    
    for row in data[2:440]:
        name = row[0].replace('*','')
        start = row[2].strftime("%Y/%m/%d")
        end = (row[2]+timedelta(days=1095)).strftime("%Y/%m/%d")
        
        first_author_count = search_pub_med(name,start,end,'Author - First')    
        author_count = search_pub_med(name,start,end,'Author')
        print([name,author_count,first_author_count])
        output_sheet.append([author_count,first_author_count])

def filterData():
    today = date.today()
    for row in output_sheet.iter_rows(values_only=True):
        if int(row[2])>0 and ((today-row[1].date())>timedelta(days=1095)):
            output_sheet.append([row[0],row[1].strftime("%m/%d/%Y"),row[2]])
            print([row[0],row[1].strftime("%Y/%m/%d"),row[2]])

filterData()
workbook.save(file_name)




